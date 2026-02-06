"""
Literaturrecherche Excel Generator
Erstellt eine vollständige Excel-Datei mit mehreren Arbeitsblättern für die Masterarbeit.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 1. Daten laden
print("📚 Lade Daten...")
screening = pd.read_csv('docs/screening_results.csv')
konzeptmatrix = pd.read_excel('docs/Konzeptmatrix_Final.xlsx')

print(f"   Screening: {len(screening)} Einträge")
print(f"   Konzeptmatrix: {len(konzeptmatrix)} Einträge")

# 2. Matching über mehrere Kriterien
# Screening hat: relevance, score, db, title, authors, year, doi, keywords
# Konzeptmatrix hat: ID, Relevanz, Score, Datenbank, Titel, Autoren, Jahr, DOI, Themen, Keywords

# Normalisiere Titel für Matching (erste 40 Zeichen, lowercase)
def normalize_title(title):
    if pd.isna(title):
        return ""
    return str(title).lower().strip()[:40]

screening['title_norm'] = screening['title'].apply(normalize_title)
konzeptmatrix['title_norm'] = konzeptmatrix['Titel'].apply(normalize_title)

# Erstelle Set der inkludierten Titel
included_titles = set(konzeptmatrix['title_norm'].dropna())

# Markiere Einträge
screening['Status'] = screening['title_norm'].apply(
    lambda x: 'Inkludiert' if x in included_titles else 'Exkludiert'
)

# Zähle
n_included = (screening['Status'] == 'Inkludiert').sum()
n_excluded = (screening['Status'] == 'Exkludiert').sum()

# Die Konzeptmatrix hat mehr Einträge als das Matching findet
# Das bedeutet, einige Papers wurden manuell aus anderen Quellen hinzugefügt
n_manual = len(konzeptmatrix) - n_included

print(f"\n📊 Ergebnis:")
print(f"   Aus Screening inkludiert: {n_included}")
print(f"   Aus Screening exkludiert: {n_excluded}")
print(f"   Manuell hinzugefügt (nicht in Screening): {n_manual}")
print(f"   Gesamt in Konzeptmatrix: {len(konzeptmatrix)}")

# 3. Statistiken erstellen
stats_data = {
    'Metrik': [
        'Gesamte Suchergebnisse (Screening)',
        'Davon inkludiert',
        'Davon exkludiert',
        'Manuell hinzugefügte Papers',
        'Finale Konzeptmatrix',
        'Inklusionsrate (%)',
        'HIGH Relevanz Papers',
        'MEDIUM Relevanz Papers',
        'Durchschnittlicher Score',
    ],
    'Wert': [
        len(screening),
        n_included,
        n_excluded,
        n_manual,
        len(konzeptmatrix),
        f"{n_included/len(screening)*100:.1f}%",
        len(konzeptmatrix[konzeptmatrix['Relevanz'] == 'HIGH']),
        len(konzeptmatrix[konzeptmatrix['Relevanz'] == 'MEDIUM']),
        f"{konzeptmatrix['Score'].mean():.1f}",
    ]
}
stats_df = pd.DataFrame(stats_data)

# Datenbank-Verteilung
db_counts = screening['db'].value_counts()
db_stats = pd.DataFrame({
    'Datenbank': db_counts.index,
    'Suchergebnisse': db_counts.values,
    'Inkludiert': [len(konzeptmatrix[konzeptmatrix['Datenbank'] == db]) for db in db_counts.index]
})
db_stats['Exkludiert'] = db_stats['Suchergebnisse'] - db_stats['Inkludiert']
db_stats['Inklusionsrate (%)'] = (db_stats['Inkludiert'] / db_stats['Suchergebnisse'] * 100).round(1)

# 4. Excel erstellen mit mehreren Blättern
print("\n📝 Erstelle Excel-Datei...")

output_path = 'docs/Literaturrecherche_Komplett.xlsx'

# Exkludierte und Inkludierte separieren
excluded = screening[screening['Status'] == 'Exkludiert'].copy()
included_from_screening = screening[screening['Status'] == 'Inkludiert'].copy()

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # Blatt 1: Übersicht/Statistiken
    stats_df.to_excel(writer, sheet_name='1_Übersicht', index=False, startrow=1)
    db_stats.to_excel(writer, sheet_name='1_Übersicht', index=False, startrow=len(stats_df) + 5)
    
    # Blatt 2: Gesamte Suche (Screening Results) - mit Status-Spalte
    screening_clean = screening.drop(columns=['title_norm'])
    screening_clean.insert(0, 'Nr', range(1, len(screening_clean) + 1))
    screening_clean.to_excel(writer, sheet_name='2_Screening_Komplett', index=False)
    
    # Blatt 3: Konzeptmatrix (finale inkludierte Papers)
    konzeptmatrix_clean = konzeptmatrix.drop(columns=['title_norm'], errors='ignore')
    konzeptmatrix_clean.to_excel(writer, sheet_name='3_Konzeptmatrix', index=False)
    
    # Blatt 4: Exkludierte Papers mit spezifischen Gründen
    excluded_clean = excluded.drop(columns=['title_norm', 'Status'])
    excluded_clean.insert(0, 'Nr', range(1, len(excluded_clean) + 1))
    
    # Spezifische Exklusionsgründe basierend auf Titel/Keywords
    def get_exclusion_reason(title):
        title_lower = str(title).lower()
        
        # Andere Industrien
        if any(x in title_lower for x in ['coal', 'mining', 'mineral']):
            return 'Andere Industrie (Bergbau)'
        if any(x in title_lower for x in ['welding', 'weld']):
            return 'Andere Industrie (Schweißen)'
        if any(x in title_lower for x in ['gas', 'natural gas', 'leakage']):
            return 'Andere Industrie (Gas/Energie)'
        if any(x in title_lower for x in ['power', 'grid', 'energy', 'smart grid']):
            return 'Andere Industrie (Energie/Netz)'
        if any(x in title_lower for x in ['construction', 'building']):
            return 'Andere Industrie (Bauwesen)'
        
        # Medizin/Gesundheit
        if any(x in title_lower for x in ['thyroid', 'medical', 'health', 'clinical', 'patient']):
            return 'Andere Domäne (Medizin)'
        
        # Andere Domänen
        if any(x in title_lower for x in ['biomimicry', 'biomimetics']):
            return 'Andere Domäne (Bionik)'
        if any(x in title_lower for x in ['tourism', 'hotel']):
            return 'Andere Domäne (Tourismus)'
        if 'robot' in title_lower:
            return 'Andere Domäne (Robotik)'
        
        # Andere spezifische Gründe
        if 'financial' in title_lower or 'sentiment' in title_lower:
            return 'Andere Domäne (Finanzen)'
        if 'fault' in title_lower and 'vehicle' not in title_lower:
            return 'Kein Automotive-Bezug'
        if 'genai tools' in title_lower or 'chatgpt' in title_lower:
            return 'Zu allgemein/keine spezifische Anwendung'
        if 'knowledge management' in title_lower:
            return 'Kein Feedback-Bezug'
        if 'prescriptive' in title_lower or 'recommender' in title_lower:
            return 'Andere Anwendung (Empfehlungssystem)'
        
        # Fallback
        return 'Thematisch nicht relevant'
    
    excluded_clean['Exklusionsgrund'] = excluded_clean['title'].apply(get_exclusion_reason)
    excluded_clean.to_excel(writer, sheet_name='4_Exkludiert', index=False)
    
    # Blatt 5: Nach Relevanz sortiert (HIGH zuerst)
    by_relevance = konzeptmatrix_clean.sort_values('Relevanz', ascending=True)
    by_relevance.to_excel(writer, sheet_name='5_Nach_Relevanz', index=False)
    
    # Blatt 6: Manuell hinzugefügte Papers (nicht im Screening)
    # Finde Papers in Konzeptmatrix die nicht im Screening gematchet wurden
    screening_titles = set(screening['title_norm'].dropna())
    manual_added = konzeptmatrix[~konzeptmatrix['title_norm'].isin(screening_titles)].copy()
    manual_added_clean = manual_added.drop(columns=['title_norm'], errors='ignore')
    manual_added_clean.insert(0, 'Nr', range(1, len(manual_added_clean) + 1))
    manual_added_clean['Herkunft'] = 'Snowballing / Manuelle Recherche'
    manual_added_clean.to_excel(writer, sheet_name='6_Manuell_Hinzugefügt', index=False)

# 6. Formatierung hinzufügen
print("🎨 Formatiere Excel...")
wb = load_workbook(output_path)

# Styles definieren
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
high_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
medium_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
excluded_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Alle Blätter formatieren
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Header formatieren
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Spaltenbreiten anpassen
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 50))
            except:
                pass
        ws.column_dimensions[column_letter].width = max(max_length + 2, 10)
    
    # Relevanz-Färbung für Konzeptmatrix und Nach_Relevanz
    if 'Konzeptmatrix' in sheet_name or 'Relevanz' in sheet_name:
        relevanz_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Relevanz':
                relevanz_col = idx
                break
        
        if relevanz_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[relevanz_col - 1].value == 'HIGH':
                    for cell in row:
                        cell.fill = high_fill
                elif row[relevanz_col - 1].value == 'MEDIUM':
                    for cell in row:
                        cell.fill = medium_fill
    
    # Exkludierte Färbung
    if 'Exkludiert' in sheet_name:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = excluded_fill

# Übersicht-Blatt speziell formatieren
ws_overview = wb['1_Übersicht']
ws_overview['A1'] = 'Literaturrecherche - Übersicht'
ws_overview['A1'].font = Font(bold=True, size=14)
ws_overview.merge_cells('A1:B1')

wb.save(output_path)
print(f"\n✅ Excel erstellt: {output_path}")
print(f"\nBlätter:")
print(f"   1. Übersicht - Statistiken und Zusammenfassung")
print(f"   2. Screening_Komplett - Alle {len(screening)} Suchergebnisse")
print(f"   3. Konzeptmatrix - {len(konzeptmatrix)} inkludierte Papers")
print(f"   4. Exkludiert - {len(excluded)} ausgeschlossene Papers")
print(f"   5. Nach_Relevanz - Sortiert HIGH → MEDIUM")
print(f"   6. Manuell_Hinzugefügt - {n_manual} Papers aus Snowballing")
