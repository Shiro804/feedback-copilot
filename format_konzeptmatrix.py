"""
Konzeptmatrix perfektionieren - Professionelle Excel-Formatierung
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList

# Excel laden
print("📂 Lade Excel...")
xl = pd.ExcelFile('docs/Literaturrecherche_Final.xlsx')
sheets_data = {name: pd.read_excel(xl, name) for name in xl.sheet_names}

# Statistiken berechnen
alle = sheets_data.get('1_Alle_Suchergebnisse', pd.DataFrame())
final = sheets_data.get('3_Final', pd.DataFrame())
exkludiert = sheets_data.get('2_Exkludiert', pd.DataFrame())

print(f"   Alle Suchergebnisse: {len(alle)}")
print(f"   Final inkludiert: {len(final)}")
print(f"   Exkludiert: {len(exkludiert)}")

# Übersicht-Daten erstellen
uebersicht_data = {
    'Metrik': [
        'Gesamte Suchergebnisse',
        'Inkludiert (Final)',
        'Exkludiert',
        'Inklusionsrate',
        '',
        'HIGH Relevanz',
        'MEDIUM Relevanz', 
        'LOW Relevanz',
        '',
        'Durchschnittlicher Score'
    ],
    'Wert': [
        len(alle),
        len(final),
        len(exkludiert),
        f"{len(final)/len(alle)*100:.1f}%" if len(alle) > 0 else "0%",
        '',
        len(final[final['relevance'] == 'HIGH']) if 'relevance' in final.columns else 0,
        len(final[final['relevance'] == 'MEDIUM']) if 'relevance' in final.columns else 0,
        len(final[final['relevance'] == 'LOW']) if 'relevance' in final.columns else 0,
        '',
        f"{final['score'].mean():.1f}" if 'score' in final.columns else 0
    ]
}
uebersicht_df = pd.DataFrame(uebersicht_data)

# Datenbank-Verteilung
if 'db' in final.columns:
    db_stats = final['db'].value_counts().reset_index()
    db_stats.columns = ['Datenbank', 'Anzahl Papers']
else:
    db_stats = pd.DataFrame({'Datenbank': [], 'Anzahl Papers': []})

# Excel neu schreiben mit allen Blättern
print("\n📝 Erstelle formatierte Excel...")
with pd.ExcelWriter('docs/Literaturrecherche_Final.xlsx', engine='openpyxl') as writer:
    # 0. Übersicht (neues Blatt)
    uebersicht_df.to_excel(writer, sheet_name='0_Übersicht', index=False, startrow=2)
    db_stats.to_excel(writer, sheet_name='0_Übersicht', index=False, startrow=len(uebersicht_df) + 5, startcol=0)
    
    # 1. Alle Suchergebnisse
    alle.to_excel(writer, sheet_name='1_Screening_Komplett', index=False)
    
    # 2. Final (Konzeptmatrix)
    final.to_excel(writer, sheet_name='2_Konzeptmatrix', index=False)
    
    # 3. Exkludiert
    exkludiert.to_excel(writer, sheet_name='3_Exkludiert', index=False)

# Formatierung
print("🎨 Formatiere...")
wb = load_workbook('docs/Literaturrecherche_Final.xlsx')

# Styles definieren
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, size=16, color='1F4E79')
subtitle_font = Font(bold=True, size=12, color='1F4E79')
high_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
medium_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
low_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Spaltenbreiten-Config
column_widths = {
    '0_Übersicht': {'A': 30, 'B': 15},
    '1_Screening_Komplett': {'A': 12, 'B': 8, 'C': 8, 'D': 65, 'E': 35, 'F': 8, 'G': 25, 'H': 45},
    '2_Konzeptmatrix': {'A': 12, 'B': 8, 'C': 8, 'D': 65, 'E': 35, 'F': 8, 'G': 25, 'H': 45},
    '3_Exkludiert': {'A': 65, 'B': 35, 'C': 8, 'D': 8, 'E': 35}
}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    if sheet_name == '0_Übersicht':
        # Titel hinzufügen
        ws['A1'] = 'Literaturrecherche - Konzeptmatrix'
        ws['A1'].font = title_font
        ws.merge_cells('A1:B1')
        
        # Untertitel für Statistiken
        ws['A2'] = 'Statistiken'
        ws['A2'].font = subtitle_font
        
        # Untertitel für Datenbanken
        row_db = len(uebersicht_df) + 5
        ws.cell(row=row_db, column=1).value = 'Verteilung nach Datenbank'
        ws.cell(row=row_db, column=1).font = subtitle_font
    else:
        # Header formatieren (Zeile 1)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Relevanz-basierte Färbung
        if 'relevance' in [c.value for c in ws[1]]:
            rel_col = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == 'relevance':
                    rel_col = idx
                    break
            
            if rel_col:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    rel_value = row[rel_col - 1].value
                    if rel_value == 'HIGH':
                        for cell in row:
                            cell.fill = high_fill
                    elif rel_value == 'MEDIUM':
                        for cell in row:
                            cell.fill = medium_fill
                    elif rel_value == 'LOW':
                        for cell in row:
                            cell.fill = low_fill
        
        # Alle Zellen formatieren
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = thin_border
        
        # Autofilter
        ws.auto_filter.ref = ws.dimensions
    
    # Spaltenbreiten setzen
    if sheet_name in column_widths:
        for col, width in column_widths[sheet_name].items():
            ws.column_dimensions[col].width = width
    
    # Zeilenhöhe für Header
    ws.row_dimensions[1].height = 25
    
    # Einfrieren der ersten Zeile
    ws.freeze_panes = 'A2'

wb.save('docs/Literaturrecherche_Final.xlsx')

print("\n✅ Fertig: docs/Literaturrecherche_Final.xlsx")
print("\nBlätter:")
print("   0. Übersicht - Statistiken und Zusammenfassung")
print(f"   1. Screening_Komplett - Alle {len(alle)} Suchergebnisse")
print(f"   2. Konzeptmatrix - {len(final)} finale Papers (HIGH/MEDIUM/LOW gefärbt)")
print(f"   3. Exkludiert - {len(exkludiert)} ausgeschlossene Papers")
